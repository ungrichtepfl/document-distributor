#!/usr/bin/env python3
"""Send bild to clients"""

from __future__ import annotations

import os
import re
import smtplib
import ssl
import string
import sys
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pprint import pprint
from typing import Mapping, Optional, Sequence

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

EMAIL_REGEX = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"


def valid_email(email: str) -> bool:
    return re.fullmatch(EMAIL_REGEX, email) is not None


def col2num(col: str) -> int:
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def name_to_mail_from_excel(excel_file_path: str,
                            first_name_column: str,
                            email_column: str,
                            last_name_column: Optional[str] = None,
                            sheet_name: Optional[str] = None,
                            start: int = 1,
                            stop: Optional[int] = None) -> Mapping[str, str]:
    workbook: Workbook = openpyxl.load_workbook(excel_file_path)
    sheet: Worksheet = workbook[sheet_name]
    stop = stop if stop is not None else sheet.max_row
    first_name_column_idx = col2num(first_name_column)
    last_name_column_idx = col2num(
        last_name_column) if last_name_column is not None else None
    email_column_idx = col2num(email_column)
    names = {}
    for row in sheet.iter_rows(min_row=start, max_row=stop):
        first_name: str = row[first_name_column_idx - 1].value
        last_name: Optional[str] = row[
            last_name_column_idx -
            1].value if last_name_column_idx is not None else None

        name = first_name.strip(" ")
        if last_name is not None:
            name = f"{name} {last_name.strip(' ')}"
        email = row[email_column_idx - 1].value.strip(" ")
        names[name] = email

    return names


def send_mail(sender_email: str,
              smtp_server: str,
              username: str,
              password: str,
              receiver_emails: Sequence[str],
              subject: str,
              message: str,
              attachment_file_paths: Sequence[str] | None = None,
              port: int = 587) -> None:
    """Send an email."""

    if attachment_file_paths is None:
        attachment_file_paths = []

    message_sent = MIMEMultipart()
    message_sent["Subject"] = subject
    message_sent["From"] = sender_email
    message_sent["To"] = ', '.join(receiver_emails)

    message_sent.attach(MIMEText(message, "plain"))

    attachment: str
    for attachment in attachment_file_paths:
        if attachment:  # Not empty string
            part = MIMEBase("application", "octet-stream")
            with open(attachment, 'rb') as bytestream:
                part.set_payload(bytestream.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {os.path.basename(attachment)}")
            message_sent.attach(part)

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_server, port) as server:
        server.ehlo()  # Can be omitted
        server.starttls(context=context)
        server.ehlo()  # Can be omitted
        server.login(username, password)
        server.sendmail(sender_email, receiver_emails,
                        message_sent.as_string())


def list_document_file_paths(document_folder_path: str) -> Sequence[str]:
    document_folder_path = os.path.abspath(
        os.path.normpath(document_folder_path))
    pdf_files = filter(lambda f: ".pdf" in f, os.listdir(document_folder_path))
    return tuple(
        map(lambda f: os.path.join(document_folder_path, f), pdf_files))


def is_name_in_document(name: str, document_path: str):
    min_matches = 2
    document_name = os.path.basename(document_path)
    total_matches = 0
    name = name.lower()
    document_name = document_name.lower()
    for token in name.split(" "):
        if token in document_name:
            total_matches += 1
            if total_matches >= min_matches:
                return True
    return False


def map_document_to_names(
        document_file_paths: Sequence[str],
        name_to_email: Mapping[str, str]) -> Mapping[str, Mapping[str, str]]:

    document_to_name_and_email = {}
    for document in document_file_paths:
        matching_names = filter(
            lambda n_e: valid_email(n_e[1]) and is_name_in_document(
                n_e[0], document), name_to_email.items())
        document_to_name_and_email[document] = dict(matching_names)

    return document_to_name_and_email


def map_name_to_documents(
    document_file_paths: Sequence[str],
    name_to_email: Mapping[str,
                           str]) -> Mapping[tuple[str, str], Sequence[str]]:

    name_to_email_and_document = {}
    for (name, email) in name_to_email.items():
        matching_documents = filter(
            lambda b: valid_email(email) and is_name_in_document(name, b),
            document_file_paths)
        name_to_email_and_document[(name, email)] = tuple(matching_documents)
    return name_to_email_and_document


def test_document_paths():
    document_file_paths = list_document_file_paths("data/rechnungen/")
    pprint(document_file_paths)


def test_email():
    msg: str = ""
    attachments: list[str] = []
    if len(sys.argv) > 1:
        msg = sys.argv[1]
    elif len(sys.argv) > 2:
        attachments = sys.argv[2:]

    sender_email: str = "info@kaesers-schloss.ch"
    smtp_server: str = "mail5.peaknetworks.net"
    username: str = "info@kaesers-schloss.ch"
    # smtp_server: str = "smtp.office365.com"
    # sender_email: str = "christoph.ungricht@outlook.com"
    # username: str = "christoph.ungricht@outlook.com"
    subject = f"Rechnung von {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    password: str = os.environ.get("EMAIL_PASSWORD")
    receiver_emails: list[str] = [
        "christoph.ungricht@outlook.com",
    ]
    assert password is not None

    if (msg or attachments):
        try:
            send_mail(sender_email=sender_email,
                      smtp_server=smtp_server,
                      username=username,
                      password=password,
                      receiver_emails=receiver_emails,
                      subject=subject,
                      message=msg,
                      attachment_file_paths=attachments)
        except Exception:
            print(f"Could not send:\n\"\n{msg}\"", file=sys.stderr)
            import traceback
            traceback.print_exc()
            return
        print(f"Sent by mail:\n\n{msg}")
        if attachments:
            attachments_txt: str = "\n".join(attachments)
            print(f"Attached documents:\n{attachments_txt}")
    else:
        print("Nothing to send. No message or attachment given.")


def test_name_to_email():
    name_to_mail = name_to_mail_from_excel(
        "./data/Zahlungen Gmüeschistli-20221019.xlsx",
        sheet_name="Abos",
        first_name_column="A",
        last_name_column="B",
        email_column="I",
        start=4,
        stop=108)
    pprint(name_to_mail)


def test_name_to_documents_and_vice_versa():
    document_file_paths = list_document_file_paths("data/rechnungen/")
    name_to_mail = name_to_mail_from_excel(
        "./data/Zahlungen Gmüeschistli-20221019.xlsx",
        sheet_name="Abos",
        first_name_column="A",
        last_name_column="B",
        email_column="I",
        start=4,
        stop=108)
    name_to_documents = map_name_to_documents(document_file_paths,
                                              name_to_mail)
    pprint(name_to_documents)
    document_to_names = map_document_to_names(document_file_paths,
                                              name_to_mail)
    pprint(document_to_names)


def app() -> int:
    """Application entry point"""

    test_email()
    test_name_to_email()
    test_document_paths()
    test_name_to_documents_and_vice_versa()

    return 0


if __name__ == "__main__":
    raise SystemExit(app())
