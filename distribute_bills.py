#!/usr/bin/env python3
"""Send bild to clients"""

from __future__ import annotations

import os
import smtplib
import ssl
import string
import sys
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pprint import pprint
from typing import Mapping, Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def col2num(col: str):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def name_to_mail_from_excel(
        excel_file_path: str,
        first_name_column: str,
        email_column: str,
        last_name_column: Optional[str] = None,
        sheet_name: Optional[str] = None,
        start: int = 1,
        stop: Optional[int] = None) -> list[Mapping[str, str]]:
    workbook: Workbook = openpyxl.load_workbook(excel_file_path)
    sheet: Worksheet = workbook[sheet_name]
    stop = stop if stop is not None else sheet.max_row
    first_name_column_idx = col2num(first_name_column)
    last_name_column_idx = col2num(
        last_name_column) if last_name_column is not None else None
    email_column_idx = col2num(email_column)
    names = {}
    for row in sheet.iter_rows(min_row=start, max_row=stop):
        first_name: str = row[first_name_column_idx - 1].value
        last_name: Optional[str] = row[
            last_name_column_idx -
            1].value if last_name_column is not None else None

        name = first_name.strip(" ")
        if last_name is not None:
            name = f"{name} {last_name.strip(' ')}"
        email = row[email_column_idx - 1].value.strip(" ")
        names[name] = email

    return names


def send_mail(sender_email: str,
              smtp_server: str,
              username: str,
              password: str,
              receiver_emails: list[str],
              subject: str,
              message: str,
              attachment_file_paths: list[str] = None,
              port: int = 587) -> None:
    """Send an email."""

    if attachment_file_paths is None:
        attachment_file_paths = []

    message_sent = MIMEMultipart()
    message_sent["Subject"] = subject
    message_sent["From"] = sender_email
    message_sent["To"] = ', '.join(receiver_emails)

    message_sent.attach(MIMEText(message, "plain"))

    attachment: str
    for attachment in attachment_file_paths:
        if attachment:  # Not empty string
            part = MIMEBase("application", "octet-stream")
            with open(attachment, 'rb') as bytestream:
                part.set_payload(bytestream.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {os.path.basename(attachment)}")
            message_sent.attach(part)

    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_server, port) as server:
        server.ehlo()  # Can be omitted
        server.starttls(context=context)
        server.ehlo()  # Can be omitted
        server.login(username, password)
        server.sendmail(sender_email, receiver_emails,
                        message_sent.as_string())


def test_email():
    msg: str = ""
    attachments: list[str] = []
    if len(sys.argv) > 1:
        msg = sys.argv[1]
    elif len(sys.argv) > 2:
        attachments = sys.argv[2:]

    sender_email: str = "info@kaesers-schloss.ch"
    smtp_server: str = "mail5.peaknetworks.net"
    username: str = "info@kaesers-schloss.ch"
    # smtp_server: str = "smtp.office365.com"
    # sender_email: str = "christoph.ungricht@outlook.com"
    # username: str = "christoph.ungricht@outlook.com"
    subject = f"Rechnung von {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    password: str = os.environ.get("EMAIL_PASSWORD")
    receiver_emails: list[str] = [
        "christoph.ungricht@outlook.com",
    ]
    assert password is not None

    if (msg or attachments):
        try:
            send_mail(sender_email=sender_email,
                      smtp_server=smtp_server,
                      username=username,
                      password=password,
                      receiver_emails=receiver_emails,
                      subject=subject,
                      message=msg,
                      attachment_file_paths=attachments)
        except Exception:
            print(f"Could not send:\n\"\n{msg}\"", file=sys.stderr)
            import traceback
            traceback.print_exc()
            return 1
        print(f"Sent by mail:\n\n{msg}")
        if attachments:
            attachments_txt: str = "\n".join(attachments)
            print(f"Attached documents:\n{attachments_txt}")
    else:
        print("Nothing to send. No message or attachment given.")


def app() -> int:
    """Application entry point"""

    test_email()

    name_to_mail = name_to_mail_from_excel(
        "./data/Zahlungen Gmüeschistli-20221019.xlsx",
        sheet_name="Abos",
        first_name_column="A",
        last_name_column="B",
        email_column="I",
        start=4,
        stop=108)
    pprint(name_to_mail)
    return 0


if __name__ == "__main__":
    raise SystemExit(app())
