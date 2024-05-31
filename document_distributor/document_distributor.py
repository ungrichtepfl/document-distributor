"""Send documents to clients"""

from __future__ import annotations

from dataclasses import dataclass
import email
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import json
import logging
import os
import re
import smtplib
import ssl
from ssl import SSLContext
import string
from typing import Iterable, Mapping, Optional, Sequence, Type

from dataclasses_json import dataclass_json
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import toolz

from document_distributor.__version__ import __version__

_log = logging.getLogger(__name__)

_EMAIL_REGEX = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
_DEFAULT_SSL_CONTEXT = ssl.create_default_context()

FileExtension = str
FilePath = str
EmailAddress = str

SUPPORTED_FILE_TYPES: Sequence[tuple[str, FileExtension]] = (("excel files", "*.xlsx"),)

_DEFAULT_MESSAGE: str = f"This message was send by ddist-{__version__}."
_DEFAULT_SUBJECT: str = f"Automatic mail from ddist-{__version__}."

MAIN_CONFIG_JSON_NAME = "general"
EMAIL_CONFIG_JSON_NAME = "email"

ROOT_PATH: str = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE_PATH = os.path.join(ROOT_PATH, "ddist-conf.json")


def dump_config(main_config: MainConfig, email_config: EmailConfig) -> None:
    obj = {
        MAIN_CONFIG_JSON_NAME: main_config.to_dict(),  # type: ignore # decorator added method
        EMAIL_CONFIG_JSON_NAME: email_config.to_dict(),  # type: ignore
    }

    if not os.path.isdir(os.path.dirname(CONFIG_FILE_PATH)):
        os.makedirs(os.path.dirname(CONFIG_FILE_PATH), exist_ok=True)

    with open(CONFIG_FILE_PATH, "w", encoding="utf-8") as f:
        f.write(json.dumps(obj))


def load_config() -> tuple[MainConfig, EmailConfig]:
    main_config: MainConfig
    email_config: EmailConfig
    if os.path.isfile(CONFIG_FILE_PATH):
        with open(CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
            obj = json.loads(f.read())

        main_config = MainConfig.from_dict(obj[MAIN_CONFIG_JSON_NAME])  # type: ignore # decorator added method
        email_config = EmailConfig.from_dict(obj[EMAIL_CONFIG_JSON_NAME])  # type: ignore
    else:
        main_config = MainConfig(document_folder_path="", excel_file_path="", first_name_column="", email_column="")
        email_config = EmailConfig(sender_email="", smtp_server="")

    return main_config, email_config


@dataclass_json
@dataclass
class MainConfig:
    document_folder_path: FilePath
    excel_file_path: FilePath
    first_name_column: str
    email_column: str
    document_file_type: Optional[list[FileExtension]] = None
    last_name_column: Optional[str] = None
    sheet_name: Optional[str] = None
    start_row_for_names: int = 1
    stop_row_for_names: Optional[int] = None


@dataclass_json
@dataclass
class EmailConfig:
    sender_email: str
    smtp_server: str
    port: int = 587
    use_starttls: bool = True
    # starttls_context: Optional[SSLContext] = _DEFAULT_SSL_CONTEXT # TODO: Add support for SSL context
    subject: str = _DEFAULT_SUBJECT
    message: str = _DEFAULT_MESSAGE
    username: Optional[str] = None
    password: Optional[str] = None


def send_emails(document_email_name: DocumentEmailName,
                email_config: EmailConfig) -> Mapping[FilePath, tuple[str, EmailAddress]]:
    document_to_email_name_not_send = {}
    config = email_config.__dict__
    for (document_path, (name, email_address)) in document_email_name.document_to_unambiguous_name_and_email().items():
        try:
            send_email(receiver_emails=[email_address], attachment_file_paths=[document_path], **config)
        except Exception as e:  # pylint: disable=broad-except, can raise a different of exceptions (user input)
            _log.error(e, exc_info=True)
            document_to_email_name_not_send[document_path] = (name, email_address)
    return document_to_email_name_not_send


@dataclass
class DocumentEmailName:
    documents_paths: Sequence[str]
    name_to_email: Mapping[str, EmailAddress]
    valid_name_to_email: Mapping[str, EmailAddress]
    valid_name_and_email_to_documents: Mapping[tuple[str, EmailAddress], Sequence[FilePath]]
    document_to_valid_names_and_emails: Mapping[FilePath, Mapping[str, EmailAddress]]

    @classmethod
    def from_data_paths(cls,
                        document_folder_path: str,
                        excel_file_path: str,
                        first_name_column: str,
                        email_column: str,
                        document_file_type: Optional[list[str]] = None,
                        last_name_column: Optional[str] = None,
                        sheet_name: Optional[str] = None,
                        start_row_for_names: int = 1,
                        stop_row_for_names: Optional[int] = None) -> DocumentEmailName:
        document_file_paths = list_document_file_paths(document_folder_path, document_file_type)
        name_to_email = name_to_mail_from_excel(excel_file_path, first_name_column, email_column, last_name_column,
                                                sheet_name, start_row_for_names, stop_row_for_names)
        valid_name_to_email = toolz.dicttoolz.valfilter(_valid_email, name_to_email)
        valid_name_and_email_to_documents = map_name_and_email_to_document(document_file_paths, valid_name_to_email)
        valid_document_to_names_and_emails = map_document_to_names_and_emails(document_file_paths, valid_name_to_email)
        return cls(document_file_paths, name_to_email, valid_name_to_email, valid_name_and_email_to_documents,
                   valid_document_to_names_and_emails)

    def document_to_unambiguous_name_and_email(self) -> Mapping[FilePath, tuple[str, EmailAddress]]:
        documents_and_one_name_to_email = toolz.valfilter(lambda v: len(v) == 1,
                                                          self.document_to_valid_names_and_emails)
        return toolz.valmap(lambda x: tuple(x.items())[0], documents_and_one_name_to_email)

    def info(self) -> str:
        names_getting_no_email: set[str] = set(
            self.name_to_email.keys()) - {name
                                          for (name, _) in self.valid_name_to_email.items()}
        names_matching_no_documents = toolz.valfilter(lambda s: len(s) == 0, self.valid_name_and_email_to_documents)

        names_getting_no_email.update(n for (n, _) in names_matching_no_documents.keys())
        document_to_name_and_email_not_sent = toolz.valfilter(lambda v: len(v) != 1,
                                                              self.document_to_valid_names_and_emails)
        document_to_unambiguous_name_and_email = self.document_to_unambiguous_name_and_email()

        info_strings = []
        if document_to_name_and_email_not_sent:
            info_strings.append(
                "The following documents could not be sent as they did not match with a name or have a invalid email:")
            info_strings.extend(f"- {os.path.basename(d):<50}" for d in document_to_name_and_email_not_sent)
        else:
            info_strings.append("All documents could be matched with a name.")

        info_strings.append("")

        if document_to_unambiguous_name_and_email:
            info_strings.append("The following documents will be sent:")
            s = (f"- {os.path.basename(d):<50} -> {n[0]} ({n[1]})"
                 for (d, n) in document_to_unambiguous_name_and_email.items())
            info_strings.extend(s)
        else:
            info_strings.append("Nothing could be sent!")
            return "\n".join(info_strings)

        info_strings.append("")

        if names_getting_no_email:
            info_strings.append("The following people will not receive any document:")
            info_strings.extend(f"- {n}" for n in names_getting_no_email)
        else:
            info_strings.append("Everybody will receive a document.")

        return "\n".join(info_strings)


def _valid_email(email_address: str) -> bool:
    return re.fullmatch(_EMAIL_REGEX, email_address) is not None


def _col2num(col: str) -> int:
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord("A")) + 1
    return num


def name_to_mail_from_excel(excel_file_path: FilePath,
                            first_name_column: str,
                            email_column: str,
                            last_name_column: Optional[str] = None,
                            sheet_name: Optional[str] = None,
                            start: int = 1,
                            stop: Optional[int] = None) -> Mapping[str, EmailAddress]:

    excel_file_path = os.path.abspath(os.path.normpath(excel_file_path))
    workbook: Workbook = openpyxl.load_workbook(excel_file_path)
    sheet: Worksheet = workbook[sheet_name] if sheet_name is not None else workbook[workbook.sheetnames[0]]
    stop = stop if stop is not None else sheet.max_row
    first_name_column_idx = _col2num(first_name_column)
    last_name_column_idx = _col2num(last_name_column) if last_name_column is not None else None
    email_column_idx = _col2num(email_column)
    names = {}
    for row in sheet.iter_rows(min_row=start, max_row=stop):
        if all(v.value is None for v in row):
            continue
        first_name_raw: Optional[str] = row[first_name_column_idx - 1].value
        first_name = first_name_raw if first_name_raw is not None else ""
        last_name: Optional[str] = row[last_name_column_idx - 1].value if last_name_column_idx is not None else None

        name = first_name.strip(" ")
        if last_name is not None:
            name = f"{name} {last_name.strip(' ')}"
        email_address_raw: Optional[str] = row[email_column_idx - 1].value
        email_address = email_address_raw.strip(" ") if email_address_raw is not None else ""
        names[name] = email_address

    return names


def send_email(
    sender_email: str,
    smtp_server: str,
    receiver_emails: Sequence[str],
    subject: str,
    message: str,
    port: int = 587,
    use_starttls: bool = True,
    starttls_context: Optional[SSLContext] = _DEFAULT_SSL_CONTEXT,
    username: Optional[str] = None,
    password: Optional[str] = None,
    attachment_file_paths: Sequence[str] | None = None,
) -> None:
    """Send an email."""
    if attachment_file_paths is None:
        attachment_file_paths = []

    message_sent = MIMEMultipart()
    message_sent["Date"] = email.utils.formatdate(localtime=True)
    message_sent["From"] = sender_email
    message_sent["To"] = ", ".join(receiver_emails)
    message_sent["Subject"] = subject

    message_sent.attach(MIMEText(message, "plain"))

    attachment: str
    for attachment in attachment_file_paths:
        if attachment:  # Not empty string
            part = _get_mime_base_from_file(attachment)
            with open(attachment, "rb") as bytestream:
                part.set_payload(bytestream.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(attachment)}"')
            message_sent.attach(part)
    smtp: Type[smtplib.SMTP] | Type[smtplib.SMTP_SSL]
    if port == smtplib.SMTP_SSL_PORT:
        # FIXME: Let user decide when to use SSL, not via port
        smtp = smtplib.SMTP_SSL
        use_starttls = False
    else:
        smtp = smtplib.SMTP

    with smtp(smtp_server, port, timeout=5) as client:
        if use_starttls:
            client.ehlo()
            client.starttls(context=starttls_context)
            # client.starttls()
        if username is not None and password is not None:
            client.ehlo()
            client.login(username, password)
        client.ehlo()
        client.sendmail(sender_email, receiver_emails, message_sent.as_string())


def _get_mime_base_from_file(file_path: str) -> MIMEBase:
    file_path_lower = file_path.lower()
    file_name = os.path.basename(file_path)
    application_type = "octet-stream"
    if file_path_lower.endswith(".pdf"):
        application_type = "pdf"

    return MIMEBase("application", application_type, name=file_name)


def list_document_file_paths(document_folder_path: FilePath,
                             file_types: Optional[Sequence[FileExtension]] = None) -> Sequence[FilePath]:
    document_folder_path = os.path.abspath(os.path.normpath(document_folder_path))
    documents: Iterable[str] = os.listdir(document_folder_path)
    if file_types is not None:
        documents = filter(lambda doc: any(ft in doc for ft in file_types), documents)  # type: ignore
    return tuple(os.path.join(document_folder_path, f) for f in documents)


def _is_name_in_document(name: str, document_path: FilePath) -> bool:
    min_matches = 2
    document_name = os.path.basename(document_path)
    total_matches = 0
    name = name.lower()
    document_name = document_name.lower()
    for token in name.split(" "):
        if token in document_name:
            total_matches += 1
            if total_matches >= min_matches:
                return True
    return False


def map_document_to_names_and_emails(
        document_file_paths: Sequence[FilePath],
        name_to_email: Mapping[str, EmailAddress]) -> Mapping[FilePath, Mapping[str, EmailAddress]]:

    document_to_name_and_email = {}
    for document in document_file_paths:
        matching_names = filter(lambda n_e: _is_name_in_document(n_e[0], document), name_to_email.items())
        document_to_name_and_email[document] = dict(matching_names)

    return document_to_name_and_email


def map_name_and_email_to_document(document_file_paths: Sequence[str],
                                   name_to_email: Mapping[str, str]) -> Mapping[tuple[str, str], Sequence[str]]:

    name_to_email_and_document = {}
    for (name, email_address) in name_to_email.items():
        matching_documents = filter(lambda b: _is_name_in_document(name, b), document_file_paths)
        name_to_email_and_document[(name, email_address)] = tuple(matching_documents)
    return name_to_email_and_document
